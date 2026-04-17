#!/usr/bin/env python3
"""Regression tests for Phase 3 P11-text youzy artifact-first consumer path."""

from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
GLOBAL_CMUX_SCRIPTS = Path("/Users/busiji/.agents/skills/cmux/scripts")
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
if str(GLOBAL_CMUX_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(GLOBAL_CMUX_SCRIPTS))

import youzy_data_replica_hook  # noqa: E402


def create_minimal_youzy_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "院校样本"
    ws.append(["院校名称", "院校特色", "院校省份", "院校类型", "院校性质", "主管部门", "排名", "难易度指数", "热度", "来源URL", "地区"])
    ws.append(["测试院校", "双一流", "北京", "综合", "公办", "教育部", 1, 95, 88, "https://example.com/c1", "华北 北京"])

    major = wb.create_sheet("专业目录")
    major.append(["一级门类", "专业类", "专业名称", "来源URL"])
    major.append(["工学", "计算机类", "计算机科学与技术", "https://example.com/m1"])

    rank = wb.create_sheet("专业热度榜")
    rank.append(["专业名称", "排名", "热度"])
    rank.append(["计算机科学与技术", 1, 99])

    wb.save(path)


def test_youzy_hook_prefers_control_packet_artifact(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        xlsx_path = temp_path / "youzy.xlsx"
        create_minimal_youzy_workbook(xlsx_path)

        payload = {
            "synced_at": "2026-04-18T03:10:00+0800",
            "youzy_college_filter_dict": [],
            "youzy_major_category_dict": [],
            "youzy_college_list": [],
            "youzy_major_list": [],
            "youzy_missing_fields": [],
        }
        control_artifact = temp_path / "youzy-control-artifact.json"
        control_artifact.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

        consumer_state = {
            "schema_version": "wb-cmux-consumer-state-v1",
            "generated_at": "2026-04-18T03:10:00+0800",
            "assignment_file": str(temp_path / "cmux-assignment.json"),
            "selected_workspace_ref": "workspace:1",
            "assignments": {
                "pm-bot": {
                    "assignment_id": "pm-001",
                    "control_packet": {
                        "artifact_path": str(control_artifact),
                    },
                }
            },
        }
        consumer_state_path = temp_path / "cmux-consumer-state-latest.json"
        consumer_state_path.write_text(json.dumps(consumer_state, ensure_ascii=False), encoding="utf-8")

        artifact_out = temp_path / "youzy-out.json"

        def fail_read_screen(*_args, **_kwargs):
            raise AssertionError("read_screen should not be called on artifact-first path")

        monkeypatch.setattr(youzy_data_replica_hook, "read_screen", fail_read_screen)
        monkeypatch.setattr(
            youzy_data_replica_hook,
            "parse_args",
            lambda: SimpleNamespace(
                workspace="workspace:1",
                surface="surface:1",
                screen_file=None,
                consumer_state_file=str(consumer_state_path),
                control_packet_artifact=str(control_artifact),
                forensic_read_pane=False,
                xlsx_path=str(xlsx_path),
                artifact_path=str(artifact_out),
            ),
        )

        assert youzy_data_replica_hook.main() == 0
        out_payload = json.loads(artifact_out.read_text(encoding="utf-8"))
        assert out_payload["youzy_missing_fields"] == []
        wb = load_workbook(xlsx_path)
        assert "youzy_college_filter_dict" in wb.sheetnames


def test_youzy_hook_blocks_screen_reads_without_forensic() -> None:
    args = SimpleNamespace(
        screen_file="/tmp/fake-screen.txt",
        surface=None,
        forensic_read_pane=False,
    )
    try:
        youzy_data_replica_hook.load_screen_text(args)
    except RuntimeError as exc:
        assert "forensic" in str(exc).lower()
    else:
        raise AssertionError("expected screen reads to be blocked without forensic flag")


def test_youzy_hook_requires_control_artifact_on_normal_path(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        xlsx_path = temp_path / "youzy.xlsx"
        create_minimal_youzy_workbook(xlsx_path)
        artifact_out = temp_path / "youzy-out.json"

        monkeypatch.setattr(
            youzy_data_replica_hook,
            "parse_args",
            lambda: SimpleNamespace(
                workspace="workspace:1",
                surface=None,
                screen_file=None,
                consumer_state_file=None,
                control_packet_artifact=None,
                forensic_read_pane=False,
                xlsx_path=str(xlsx_path),
                artifact_path=str(artifact_out),
            ),
        )
        try:
            youzy_data_replica_hook.main()
        except RuntimeError as exc:
            assert "control-packet artifact is required" in str(exc)
        else:
            raise AssertionError("expected normal path to require control-packet artifact")
